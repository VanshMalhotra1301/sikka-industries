import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_stock_excel(products_list):
    """
    Constructs an isolated, beautifully styled Excel sheet listing active inventory blocks.
    Returns a raw BytesIO buffer containing the file stream.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse Stock Audit"

    # Define strict professional typography scales and surface styling fills
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    alert_font = Font(name=font_family, size=10, bold=True, color="EF4444")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Set Column Metadata Headers Layout Matrix Rows
    headers = ["Component Code", "Descriptive Name", "Classification", "Cost Unit (₹)", "Selling Value (₹)", "Current Qty", "Status"]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Map target application dataset items
    for p in products_list:
        is_low = p.stock_quantity <= p.low_stock_threshold
        status_text = "LOW ALERT" if is_low else "HEALTHY"
        
        row_data = [
            p.code,
            p.name,
            p.category,
            p.purchase_cost,
            p.selling_price if p.category == 'Finished Goods' else 0.0,
            p.stock_quantity,
            status_text
        ]
        ws.append(row_data)
        
        # Style trailing rows dynamically safely
        current_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = alert_font if (is_low and col_idx == 6) else data_font
            cell.border = thin_border
            
            # Apply currency or numerical formatting setups where applicable
            if col_idx in [4, 5]:
                cell.number_format = '"₹"#,##0.00'
            elif col_idx == 6:
                cell.number_format = '#,##0.00'

    # Auto-adjust explicit tracking cell spacing layout gaps
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

